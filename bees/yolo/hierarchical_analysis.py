"""
Hierarchical Analysis Pipeline for Nested Folder Structures

Final Structure:
  Вид_А/
    Control/              ← Проба-контроль
      Сэмпл_1/            ← Повторность 1 (серия фото)
        photo1.jpg
        photo2.jpg
      Сэмпл_2/            ← Повторность 2
        photo1.jpg
        photo2.jpg
    Проба_1/              ← Проба-опыт
      Сэмпл_1/            ← Повторность 1
        photo1.jpg
        photo2.jpg
      Сэмпл_2/            ← Повторность 2
        photo1.jpg
        photo2.jpg
    Проба_2/
      Сэмпл_1/
      Сэмпл_2/
  Вид_Б/
    Control/
    Проба_1/
    ...

Logic:
  1. Each "sample" folder = one repetition with series of photos
  2. Titer per sample: Σspores / (4 * (S_photo/780²) * N_photos)
  3. Titer per probe = mean of sample titers (repetitions)
  4. p-value per probe = two-sample Welch's t-test: sample titers vs control sample titers
"""

import logging
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
import numpy as np

from ..titer import TiterCalculator, create_calculator_from_config
from ..config_loader import create_config_manager
try:
    from .sahi_inference import SAHIDetector
    SAHI_AVAILABLE = True
except ImportError:
    SAHI_AVAILABLE = False

logger = logging.getLogger(__name__)


@dataclass
class SampleResult:
    """Results for a single sample folder (one repetition)."""
    sample_name: str
    probe_name: str
    type_name: str
    photo_data: List[Tuple[int, int, int]]  # [(count, width, height), ...]
    titer: float = 0.0
    breakdown: dict = field(default_factory=dict)

    @property
    def n_photos(self) -> int:
        return len(self.photo_data)

    @property
    def total_spores(self) -> int:
        return sum(count for count, _, _ in self.photo_data)


@dataclass
class ProbeResult:
    """Results for a probe (row) with multiple samples (repetitions)."""
    probe_name: str
    type_name: str
    samples: List[SampleResult]
    is_control: bool = False

    @property
    def n_samples(self) -> int:
        return len(self.samples)

    @property
    def total_photos(self) -> int:
        return sum(s.n_photos for s in self.samples)

    @property
    def total_spores(self) -> int:
        return sum(s.total_spores for s in self.samples)

    @property
    def sample_titers(self) -> List[float]:
        return [s.titer for s in self.samples]

    @property
    def mean_titer(self) -> float:
        if not self.samples:
            return 0.0
        return float(np.mean(self.sample_titers))

    @property
    def std_titer(self) -> float:
        if len(self.samples) < 2:
            return 0.0
        return float(np.std(self.sample_titers, ddof=1))


@dataclass
class TypeStatistics:
    """Statistics for a type (sheet) with multiple probes (rows)."""
    type_name: str
    probes: List[ProbeResult]
    control_probe: Optional[ProbeResult] = None

    @property
    def control_titers(self) -> List[float]:
        if self.control_probe is None:
            return []
        return self.control_probe.sample_titers

    @property
    def control_mean(self) -> Optional[float]:
        if self.control_probe is None:
            return None
        return self.control_probe.mean_titer

    def get_probe_p_value(self, probe: ProbeResult) -> Optional[float]:
        """
        Two-sample Welch's t-test: probe sample titers vs control sample titers.
        Returns None if insufficient data (< 2 samples in either group).
        """
        if self.control_probe is None:
            return None

        control_titers = self.control_titers
        probe_titers = probe.sample_titers

        if len(control_titers) < 2 or len(probe_titers) < 2:
            return None

        calc = TiterCalculator()
        _, p_value = calc.two_sample_ttest(probe_titers, control_titers)
        return float(p_value)


class HierarchicalAnalyzer:
    """Analyzes spore images in nested folder structures."""

    def __init__(self, config_path: str = "config.yaml", use_yolo: bool = True,
                 weights_path: Optional[str] = None, use_sahi: bool = False):
        self.config_manager = create_config_manager(config_path)
        self.use_yolo = use_yolo
        self.weights_path = weights_path
        # SAHI параметры
        self.weights_path = weights_path
        self.use_sahi = use_sahi and SAHI_AVAILABLE
        self.detector = None
        self.titer_calculator = create_calculator_from_config(self.config_manager)
        self._initialize_detector()

    def _initialize_detector(self):
        if not self.use_yolo:
            from bees.spore_analysis_pipeline import SporeAnalysisPipeline
            self.pipeline = SporeAnalysisPipeline(
                self.config_manager, None, None, use_yolo=False
            )
            logger.info("✓ OpenCV detector initialized")
            return

        # --- YOLO ветка ---
        if self.use_sahi:
            if not SAHI_AVAILABLE:
                raise ImportError("SAHI not installed. Run: pip install sahi")

            # Читаем настройки из секции tiley.predict конфига
            tiley_cfg = self.config_manager.get_tiley()
            predict_cfg = tiley_cfg.get("predict", {})

            # Путь к весам: приоритет у --weights, иначе из config, иначе стандартный
            if self.weights_path:
                model_path = self.weights_path
            else:
                weights_from_cfg = predict_cfg.get("weights")
                if weights_from_cfg:
                    model_path = weights_from_cfg
                else:
                    from bees.yolo import YOLOConfig
                    yolo_config = YOLOConfig.from_config_manager(self.config_manager)
                    model_path = str(yolo_config.get_trained_model_path())

            conf = predict_cfg.get("conf", self.config_manager.get_float_param('yolo_confidence', 0.25))
            device = predict_cfg.get("device", "cuda:0")
            model_type = predict_cfg.get("model_type", "ultralytics")
            self.slice_height = predict_cfg.get("tile_size", 1024)
            self.slice_width = self.slice_height
            self.overlap_height_ratio = predict_cfg.get("overlap", 0.2)
            self.overlap_width_ratio = self.overlap_height_ratio

            self.detector = SAHIDetector(
                model_path=model_path,
                model_type=model_type,
                confidence_threshold=conf,
                device=device
            )
            logger.info(f"✓ SAHI detector initialized (tile={self.slice_height}, overlap={self.overlap_height_ratio})")
        else:
            from bees.yolo import YOLOConfig, SporeDetector
            yolo_config = YOLOConfig.from_config_manager(self.config_manager)
            self.detector = SporeDetector(yolo_config)
            if self.weights_path:
                self.detector.load_weights(self.weights_path)
            logger.info("✓ Standard YOLO detector initialized")

    def is_control_name(self, name: str) -> bool:
        control_names = {'control', 'контроль', 'ctrl', 'кontроль', 'контр'}
        return name.lower().strip() in control_names

    def find_images(self, folder: Path) -> List[Path]:
        extensions = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff'}
        images = []
        for ext in extensions:
            images.extend(folder.glob(f"*{ext}"))
        # Удаляем возможные дубликаты (на случай, если в папке есть файлы с разным регистром)
        unique = {}
        for p in images:
            key = str(p.resolve()).lower()
            if key not in unique:
                unique[key] = p
        return sorted(unique.values())

    def analyze_sample_folder(self, sample_path: Path, probe_name: str, type_name: str) -> SampleResult:
        """Analyze all images in a sample folder (one repetition)."""
        images = self.find_images(sample_path)

        if not images:
            logger.warning(f"No images found in {sample_path}")
            return SampleResult(
                sample_name=sample_path.name,
                probe_name=probe_name,
                type_name=type_name,
                photo_data=[]
            )

        photo_data: List[Tuple[int, int, int]] = []

        for image_path in images:
            try:
                from PIL import Image
                with Image.open(image_path) as img:
                    img_width, img_height = img.size

                if self.use_yolo:
                    if self.use_sahi:
                        import cv2
                        img_bgr = cv2.imread(str(image_path))
                        if img_bgr is None:
                            logger.error(f"Cannot read {image_path}")
                            continue
                        detections = self.detector.detect_sliced(
                            img_bgr,
                            slice_height=self.slice_height,
                            slice_width=self.slice_width,
                            overlap_height_ratio=self.overlap_height_ratio,
                            overlap_width_ratio=self.overlap_width_ratio
                        )
                        count = len(detections)
                    else:
                        detections = self.detector.detect(str(image_path))
                        count = len(detections)

                photo_data.append((count, img_width, img_height))

                logger.debug(
                    f"      {image_path.name}: {count} spores, size={img_width}x{img_height}"
                )

            except Exception as e:
                logger.error(f"Failed to analyze {image_path}: {e}")
                continue

        sample_result = SampleResult(
            sample_name=sample_path.name,
            probe_name=probe_name,
            type_name=type_name,
            photo_data=photo_data
        )

        if photo_data:
            sample_result.titer = self.titer_calculator.calculate_sample_titer(photo_data)
            _, sample_result.breakdown = self.titer_calculator.calculate_sample_titer_safe(photo_data)

        return sample_result

    def analyze_probe(self, probe_path: Path, type_name: str) -> ProbeResult:
        """Analyze all samples (repetitions) within a probe."""
        logger.info(f"  Analyzing probe: {probe_path.name}")

        sample_results: List[SampleResult] = []

        for item in sorted(probe_path.iterdir()):
            if not item.is_dir():
                continue

            logger.info(f"    Analyzing sample: {item.name}")
            sample_result = self.analyze_sample_folder(item, probe_path.name, type_name)
            sample_results.append(sample_result)

            logger.info(
                f"      ✓ {item.name}: {sample_result.total_spores} spores in "
                f"{sample_result.n_photos} photos, titer={sample_result.titer:.4f}"
            )

        is_control = self.is_control_name(probe_path.name)
        probe_result = ProbeResult(
            probe_name=probe_path.name,
            type_name=type_name,
            samples=sample_results,
            is_control=is_control
        )

        logger.info(
            f"    ✓ Probe {probe_path.name}: {probe_result.n_samples} samples, "
            f"mean_titer={probe_result.mean_titer:.4f} ± {probe_result.std_titer:.4f}"
        )

        return probe_result

    def analyze_type(self, type_path: Path) -> TypeStatistics:
        """Analyze one 'Вид' (type) folder."""
        logger.info(f"Analyzing type: {type_path.name}")

        probe_results: List[ProbeResult] = []
        control_probe: Optional[ProbeResult] = None

        for item in sorted(type_path.iterdir()):
            if not item.is_dir():
                continue

            probe_result = self.analyze_probe(item, type_path.name)
            probe_results.append(probe_result)

            if probe_result.is_control:
                control_probe = probe_result
                logger.info(f"  ✓ CONTROL probe: {probe_result.probe_name}")

        type_stats = TypeStatistics(
            type_name=type_path.name,
            probes=probe_results,
            control_probe=control_probe
        )

        # Исправленная строка:
        control_val = f"{type_stats.control_mean:.4f}" if type_stats.control_mean is not None else "None"
        logger.info(f"✓ Type {type_path.name}: {len(probe_results)} probes, control={control_val}")

        return type_stats

    def analyze_hierarchy(self, root_dir: Path) -> Dict[str, TypeStatistics]:
        """Analyze complete folder hierarchy."""
        logger.info(f"Starting hierarchical analysis of {root_dir}")

        results: Dict[str, TypeStatistics] = {}

        for item in sorted(root_dir.iterdir()):
            if not item.is_dir():
                continue

            type_stats = self.analyze_type(item)
            results[item.name] = type_stats

        logger.info(f"✓ Analysis complete: {len(results)} types")
        return results

    def generate_excel_report(self, analysis_results, output_dir, filename="hierarchical_analysis.xlsx"):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        output_path = output_dir / filename
        wb = Workbook()

        for type_name, type_stats in sorted(analysis_results.items()):
            sheet_name = type_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Title
            ws.append([f"Вид (Type): {type_name}"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            title_cell = ws.cell(row=1, column=1)
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.append([])

            # Headers
            headers = ['Probe', 'n', 'Titer ± Std', 'p-value']
            ws.append(headers)

            # Header style
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for col_idx in range(1, 5):
                cell = ws.cell(row=3, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            row_idx = 4
            for probe in sorted(type_stats.probes, key=lambda p: p.probe_name):
                p_value = type_stats.get_probe_p_value(probe)
                p_val_str = f"{p_value:.6f}" if p_value is not None else "N/A"
                mean_std_str = f"{probe.mean_titer:.4f} ± {probe.std_titer:.6f}"

                ws.append([
                    probe.probe_name,
                    probe.n_samples,
                    mean_std_str,
                    p_val_str
                ])

                if not probe.is_control and p_value is not None:
                    p_cell = ws.cell(row=row_idx, column=4)
                    if p_value < 0.05:
                        p_cell.font = Font(color="006100", bold=True)
                        p_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        p_cell.font = Font(color="9C0006")
                        p_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                row_idx += 1

            column_widths = [25, 15, 22, 22]
            for col_idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            ws.row_dimensions[3].height = 30

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=4):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 3:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        output_dir.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"✓ Excel report saved: {output_path}")
        return output_path
    def generate_markdown_summary(
        self,
        analysis_results: Dict[str, TypeStatistics],
        output_dir: Path,
        filename: str = "analysis_summary.md"
    ) -> Path:
        """Generate Markdown summary report."""
        output_path = output_dir / filename

        content = "# Hierarchical Spore Analysis Report\n\n"

        for type_name, type_stats in sorted(analysis_results.items()):
            content += f"## {type_name}\n\n"
            control_str = f"{type_stats.control_mean:.4f}" if type_stats.control_mean is not None else "N/A"
            content += f"**Control:** {control_str}\n\n"
            content += "| Probe | n | Titer | Std Dev | p-value | Control |\n"
            content += "|-------|---|------|---------|---------|---------|\n"

            for probe in sorted(type_stats.probes, key=lambda p: p.probe_name):
                p_value = type_stats.get_probe_p_value(probe)
                p_str = f"{p_value:.6f}" if p_value is not None else "N/A"
                ctrl_mark = "✓" if probe.is_control else ""
                content += (
                    f"| {probe.probe_name} | {probe.n_samples} | "
                    f"{probe.mean_titer:.4f} | {probe.std_titer:.6f} | "
                    f"{p_str} | {ctrl_mark} |\n"
                )

            content += "\n"

        output_path.write_text(content, encoding='utf-8')
        logger.info(f"✓ Markdown summary saved: {output_path}")
        return output_path

    def run_complete_analysis(
        self,
        root_dir: Path,
        output_dir: Path,
        excel_filename: str = "hierarchical_analysis.xlsx",
        markdown_filename: str = "analysis_summary.md"
    ) -> Dict[str, Path]:
        """Run complete analysis pipeline."""
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        results = self.analyze_hierarchy(root_dir)

        if not results:
            logger.error("No analysis results generated")
            return {}

        reports = {}

        excel_path = self.generate_excel_report(results, output_dir, excel_filename)
        reports['excel'] = excel_path

        md_path = self.generate_markdown_summary(results, output_dir, markdown_filename)
        reports['markdown'] = md_path

        logger.info("✓ All reports generated successfully")
        return reports


def main():
    """Command-line interface for hierarchical analysis."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Hierarchical Analysis: Вид/Проба/Сэмпл(фото)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Structure:
  input_dir/
    Вид_А/
      Control/
        Сэмпл_1/          ← Repetition 1
          photo1.jpg
        Сэмпл_2/          ← Repetition 2
          photo1.jpg
      Проба_1/
        Сэмпл_1/
        Сэмпл_2/

Examples:
  python -m bees.hierarchical_analysis --input-dir ./data --output-dir ./results
  python -m bees.hierarchical_analysis --input-dir ./data --no-yolo
        """
    )

    parser.add_argument("--input-dir", type=Path, required=True,
                        help="Root directory with Type/Probe/Sample(photos) structure")
    parser.add_argument("--output-dir", type=Path, default=Path("hierarchical_output"),
                        help="Output directory (default: hierarchical_output)")
    parser.add_argument("--config", type=str, default="config.yaml",
                        help="Config file (default: config.yaml)")
    parser.add_argument("--no-yolo", action="store_true",
                        help="Use OpenCV instead of YOLO")
    parser.add_argument("--use-sahi", action="store_true",
                        help="Use SAHI sliced inference (reads tiling params from config.yaml)")
    parser.add_argument("--weights", type=str, default=None,
                        help="Path to custom YOLO weights (.pt file)")  # <-- ЭТУ СТРОКУ
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

    try:
        if not args.input_dir.exists():
            logger.error(f"Input directory not found: {args.input_dir}")
            return 1

        analyzer = HierarchicalAnalyzer(
            config_path=args.config,
            use_yolo=not args.no_yolo,
            weights_path=args.weights,
            use_sahi=args.use_sahi
        )
        reports = analyzer.run_complete_analysis(root_dir=args.input_dir, output_dir=args.output_dir)

        logger.info("✓ Analysis completed!")
        for report_type, report_path in reports.items():
            logger.info(f"  - {report_type}: {report_path}")

        return 0

    except Exception as e:
        logger.error(f"Analysis failed: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
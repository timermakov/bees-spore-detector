"""
Reporting module for bee spore analysis.

This module provides functionality for generating reports in various formats
including Markdown and Excel.
"""

import os
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Union
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from bees.titer import TiterCalculator

# Configure logging
logger = logging.getLogger(__name__)


class MarkdownReporter:
    """Handles generation of Markdown reports."""
    
    def __init__(self, output_dir: Union[str, Path]):
        """
        Initialize the Markdown reporter.
        
        Args:
            output_dir: Directory for output reports
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def write_image_report(self, 
                          image_path: Union[str, Path], 
                          count: int, 
                          titer_value: float,
                          output_filename: Optional[str] = None) -> Path:
        """
        Write a Markdown report for a single image.
        
        Args:
            image_path: Path to the analyzed image
            count: Number of detected spores
            titer_value: Calculated titer value
            output_filename: Optional custom filename for the report
            
        Returns:
            Path to the generated report file
            
        Example:
            >>> reporter = MarkdownReporter("results/")
            >>> report_path = reporter.write_image_report("image.jpg", 150, 12.5)
            >>> print(f"Report saved to: {report_path}")
        """
        image_name = Path(image_path).name
        
        if output_filename is None:
            base_name = Path(image_name).stem
            output_filename = f"{base_name}.md"
        
        output_path = self.output_dir / output_filename
        
        content = self._generate_image_report_content(image_name, count, titer_value)
        
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            logger.debug(f"Markdown report saved: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to write Markdown report {output_path}: {e}")
            raise
    
    def _generate_image_report_content(self, 
                                     image_name: str, 
                                     count: int, 
                                     titer_value: float) -> str:
        """Generate the content for an image report."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        return f"""# Bee Spore Analysis Report

**Image:** {image_name}  
**Analysis Date:** {timestamp}

## Results

- **Spore Count:** {count}
- **Titer:** {titer_value:.2f} million spores/ml

## Analysis Parameters

- **Method:** Goryaev Chamber
- **Detection Algorithm:** Computer Vision with Ellipse Fitting

"""
    
    def write_group_report(self, 
                          group_prefix: str, 
                          image_counts: List[int], 
                          group_titer: float) -> Path:
        """
        Write a Markdown report for a group of images.
        
        Args:
            group_prefix: Group prefix name
            image_counts: List of spore counts for each image
            group_titer: Calculated titer for the group
            
        Returns:
            Path to the generated report file
        """
        output_filename = f"{group_prefix}_group_report.md"
        output_path = self.output_dir / output_filename
        
        content = self._generate_group_report_content(group_prefix, image_counts, group_titer)
        
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            logger.debug(f"Group report saved: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to write group report {output_path}: {e}")
            raise
    
    def _generate_group_report_content(self, 
                                     group_prefix: str, 
                                     image_counts: List[int], 
                                     group_titer: float) -> str:
        """Generate the content for a group report."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        content = f"""# Group Analysis Report

**Group:** {group_prefix}  
**Analysis Date:** {timestamp}

## Individual Results

"""
        
        for i, count in enumerate(image_counts, 1):
            content += f"- **Sample {i}:** {count} spores\n"
        
        content += f"""
## Group Summary

- **Total Spores:** {sum(image_counts)}
- **Group Titer:** {group_titer:.2f} million spores/ml
- **Sample Count:** {len(image_counts)}

## Analysis Parameters

- **Method:** Goryaev Chamber
- **Grouping:** Triplicate Analysis
"""
        
        return content


class ExcelReporter:
    """Handles generation of Excel reports."""
    
    def __init__(self, output_dir: Union[str, Path]):
        """
        Initialize the Excel reporter.
        
        Args:
            output_dir: Directory for output reports
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.titer_calculator = TiterCalculator()
    
    def export_results(self, 
                      groups_results: Dict[str, List[Tuple[int, float]]], 
                      output_filename: str = "spore_analysis_report.xlsx") -> Path:
        """
        Export analysis results to Excel format.
        
        Args:
            groups_results: Dictionary mapping group prefixes to lists of (count, titer) tuples
            output_filename: Name of the output Excel file
            
        Returns:
            Path to the generated Excel file
            
        Example:
            >>> reporter = ExcelReporter("results/")
            >>> excel_path = reporter.export_results(groups_data)
            >>> print(f"Excel report saved to: {excel_path}")
        """
        output_path = self.output_dir / output_filename
        
        try:
            workbook = self._create_workbook(groups_results)
            workbook.save(output_path)
            
            logger.info(f"Excel report saved: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to create Excel report {output_path}: {e}")
            raise
    
    def _create_workbook(self, groups_results: Dict[str, List[Tuple[int, float]]]) -> Workbook:
        """Create the Excel workbook with formatted data."""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Spore Analysis Report'
        
        # Apply styling
        self._apply_workbook_styling(ws)
        
        # Add header
        self._add_header(ws)
        
        # Add data
        self._add_data(ws, groups_results)
        
        # Format columns
        self._format_columns(ws)
        
        return wb
    
    def _apply_workbook_styling(self, worksheet) -> None:
        """Apply basic styling to the worksheet."""
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header styling
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    def _add_header(self, worksheet) -> None:
        """Add column headers to the worksheet."""
        headers = ["Sample Group", "Sample Number", "Spore Count", "Group Titer (million spores/ml)"]
        worksheet.append(headers)
    
    def _add_data(self, worksheet, groups_results: Dict[str, List[Tuple[int, float]]]) -> None:
        """Add data rows to the worksheet."""
        start_row = 2
        
        for group_prefix, rows in groups_results.items():
            # Calculate group titer
            group_counts = [count for count, _ in rows]
            group_titer = self.titer_calculator.calculate_titer(group_counts)
            
            # Add data rows
            for idx, (count, _) in enumerate(rows, start=1):
                worksheet.append([group_prefix, idx, count, group_titer])
            
            # Merge cells for group name and titer
            end_row = start_row + len(rows) - 1
            if end_row >= start_row:
                self._merge_group_cells(worksheet, start_row, end_row)
            
            start_row = end_row + 1
    
    def _merge_group_cells(self, worksheet, start_row: int, end_row: int) -> None:
        """Merge cells for group name and titer columns."""
        # Merge group name cells
        worksheet.merge_cells(
            start_row=start_row, 
            start_column=1, 
            end_row=end_row, 
            end_column=1
        )
        
        # Merge titer cells
        worksheet.merge_cells(
            start_row=start_row, 
            start_column=4, 
            end_row=end_row, 
            end_column=4
        )
        
        # Center align merged cells
        for col in [1, 4]:
            cell = worksheet.cell(row=start_row, column=col)
            cell.alignment = Alignment(vertical='center', horizontal='center')
    
    def _format_columns(self, worksheet) -> None:
        """Format column widths and apply borders."""
        # Set column widths
        column_widths = [25, 15, 20, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Apply borders to all cells with data
        self._apply_borders(worksheet)
    
    def _apply_borders(self, worksheet) -> None:
        """Apply borders to all cells with data."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border


class ReportManager:
    """Manages the generation of multiple report types."""
    
    def __init__(self, output_dir: Union[str, Path]):
        """
        Initialize the report manager.
        
        Args:
            output_dir: Directory for output reports
        """
        self.output_dir = Path(output_dir)
        self.titer_calculator = TiterCalculator()
        self.markdown_reporter = MarkdownReporter(output_dir)
        self.excel_reporter = ExcelReporter(output_dir)
    
    def generate_all_reports(self, 
                           groups_results: Dict[str, List[Tuple[int, float]]],
                           image_results: Optional[Dict[str, dict]] = None) -> dict:
        """
        Generate all types of reports.
        
        Args:
            groups_results: Dictionary of grouped results
            image_results: Optional dictionary of individual image results
            
        Returns:
            Dictionary with paths to generated reports
        """
        reports = {}
        
        try:
            # Generate Excel report
            excel_path = self.excel_reporter.export_results(groups_results)
            reports['excel'] = excel_path
            
            # Generate individual image reports if provided
            if image_results:
                image_reports = {}
                for image_path, result in image_results.items():
                    report_path = self.markdown_reporter.write_image_report(
                        image_path, 
                        result['count'], 
                        result['titer']
                    )
                    image_reports[image_path] = report_path
                reports['individual_reports'] = image_reports
            
            # Generate group reports
            group_reports = {}
            for prefix, rows in groups_results.items():
                counts = [count for count, _ in rows]
                group_titer = self.titer_calculator.calculate_titer(counts)
                report_path = self.markdown_reporter.write_group_report(
                    prefix, counts, group_titer
                )
                group_reports[prefix] = report_path
            reports['group_reports'] = group_reports
            
            logger.info(f"Generated {len(reports)} report types")
            return reports
            
        except Exception as e:
            logger.error(f"Failed to generate reports: {e}")
            raise





